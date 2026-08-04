#!/usr/bin/env python3
"""
PostgreSQL Attendance Exporter to Excel (.xlsx) & CSV

This script connects to the PostgreSQL database, extracts all attendance
records, and organizes them by session.

Dependencies:
    pip install psycopg2-binary pandas openpyxl python-dotenv

Usage:
    python3 export_sessions.py [options]
"""

import sys
import os
import csv
import argparse
from datetime import datetime

# Load environment variables robustly
try:
    from dotenv import load_dotenv, find_dotenv
    # Find .env file starting from the current working directory or script directory
    env_path = find_dotenv(usecwd=True)
    if env_path:
        load_dotenv(env_path)
        print(f"[*] Loaded environment variables from {env_path}")
    else:
        print("[*] No .env file found. Using system environment variables or defaults.")
except ImportError:
    print("[!] 'python-dotenv' is missing. Cannot load .env file automatically.")
    pass

# PostgreSQL driver
try:
    import psycopg2
    import psycopg2.extras
except ImportError:
    print("Error: 'psycopg2-binary' library is missing.")
    print("Please install dependencies by running:")
    print("    pip install psycopg2-binary pandas openpyxl python-dotenv")
    sys.exit(1)

# Optional Pandas & OpenPyXL for styled Excel exports
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def construct_database_url():
    """
    Constructs the Postgres connection string from individual environment
    variables if DATABASE_URL is not explicitly defined.
    """
    # 1. If full URL is provided, use it directly
    url = os.getenv("DATABASE_URL")
    if url:
        return url

    # 2. Otherwise build from components
    user = os.getenv("PG_USER", "postgres")
    password = os.getenv("PG_PASSWORD", "postgres")
    host = os.getenv("PG_HOST", "localhost")
    port = os.getenv("PG_PORT", "5432")
    db = os.getenv("PG_DATABASE", "attendance_geotag")

    return f"postgres://{user}:{password}@{host}:{port}/{db}"


def fetch_data(conn):
    """
    Fetch sessions, attendances, locations, and batches from Postgres.
    """
    print("[*] Fetching sessions and attendance records...")

    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    cur.execute("SELECT id, location_id, batch_id, created_at FROM sessions")
    sessions_raw = cur.fetchall()

    cur.execute(
        """
        SELECT session_id, student_name, roll_number, captured_at, verified,
               flagged, flag_reason, face_detected, webauthn_verified,
               distance_from_location, ip_address
        FROM attendances
        """
    )
    attendances_raw = cur.fetchall()

    print(f"[*] Retrieved {len(sessions_raw)} sessions and {len(attendances_raw)} attendance records.")

    # Fetch lookup data for enrichment
    cur.execute("SELECT id, name FROM locations")
    locations_map = {str(row["id"]): row["name"] for row in cur.fetchall()}

    cur.execute("SELECT id, name FROM batches")
    batches_map = {str(row["id"]): row["name"] for row in cur.fetchall()}

    cur.close()

    # Organize sessions
    sessions = {}
    for s in sessions_raw:
        s_id = str(s["id"])
        loc_id = str(s["location_id"]) if s["location_id"] else ""
        batch_id = str(s["batch_id"]) if s["batch_id"] else ""

        created_at = s["created_at"]

        sessions[s_id] = {
            "session_id": s_id,
            "location_name": locations_map.get(loc_id, loc_id),
            "batch_name": batches_map.get(batch_id, batch_id),
            "date": created_at.strftime("%Y-%m-%d") if isinstance(created_at, datetime) else str(created_at)
        }

    # Process attendances
    attendances_by_session = {}
    all_attendance_rows = []

    for a in attendances_raw:
        sess_id = str(a["session_id"])

        captured_at = a["captured_at"]
        captured_at_str = captured_at.strftime("%Y-%m-%d %H:%M:%S") if isinstance(captured_at, datetime) else str(captured_at)

        row = {
            "Session ID": sess_id,
            "Batch": sessions.get(sess_id, {}).get("batch_name", "Unknown"),
            "Location": sessions.get(sess_id, {}).get("location_name", "Unknown"),
            "Session Date": sessions.get(sess_id, {}).get("date", "Unknown"),
            "Student Name": a["student_name"] or "N/A",
            "Roll Number": a["roll_number"] or "N/A",
            "Time": captured_at_str,
            "Verified": a["verified"] or False,
            "Flagged": a["flagged"] or False,
            "Flag Reason": a["flag_reason"] or "",
            "Face Detected": a["face_detected"] if a["face_detected"] is not None else False,
            "WebAuthn Verified": a["webauthn_verified"] if a["webauthn_verified"] is not None else False,
            "Distance (m)": round(a["distance_from_location"] or 0.0, 2),
            "IP Address": a["ip_address"] or ""
        }

        all_attendance_rows.append(row)

        if sess_id not in attendances_by_session:
            attendances_by_session[sess_id] = []
        attendances_by_session[sess_id].append(row)

    return sessions, attendances_by_session, all_attendance_rows


def export_to_excel(sessions, attendances_by_session, output_path):
    """
    Export to Excel where each session gets its own formatted sheet.
    """
    if not HAS_OPENPYXL or not HAS_PANDAS:
        print("[!] Pandas or OpenPyXL missing, cannot create Excel file. Use --format csv instead.")
        return

    wb = openpyxl.Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    alt_fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    if not sessions:
        ws = wb.create_sheet(title="No Data")
        ws.cell(1, 1, "No sessions found in the database.")
        wb.save(output_path)
        return

    for idx, (sess_id, session_info) in enumerate(sessions.items()):
        # Create a safe sheet name (max 31 chars, no invalid chars)
        batch = session_info["batch_name"]
        date = session_info["date"]
        # E.g. "BTech_CS_2026-07-31" or fallback to session ID
        sheet_title = f"{batch}_{date}"
        sheet_title = "".join(c for c in sheet_title if c.isalnum() or c in (" ", "_", "-"))
        sheet_title = sheet_title[:31] if sheet_title else sess_id[:31]

        # Ensure unique sheet names
        original_title = sheet_title
        counter = 1
        while sheet_title in wb.sheetnames:
            suffix = f"_{counter}"
            sheet_title = original_title[:31 - len(suffix)] + suffix
            counter += 1

        ws = wb.create_sheet(title=sheet_title)

        # Get attendances for this session
        sess_attendances = attendances_by_session.get(sess_id, [])

        # We drop the redundant Session columns for the individual sheets
        display_rows = []
        for a in sess_attendances:
            display_rows.append({
                "Student Name": a["Student Name"],
                "Roll Number": a["Roll Number"],
                "Time": a["Time"],
                "Verified": a["Verified"],
                "Flagged": a["Flagged"],
                "Flag Reason": a["Flag Reason"],
                "Face Detected": a["Face Detected"],
                "WebAuthn": a["WebAuthn Verified"],
                "Distance (m)": a["Distance (m)"]
            })

        if not display_rows:
            ws.cell(1, 1, "No attendance records for this session.")
            continue

        df = pd.DataFrame(display_rows)

        # Write header
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[1].height = 24

        # Write data
        for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            ws.row_dimensions[row_idx].height = 18
            use_alt = (row_idx % 2 == 0)
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.border = thin_border
                if use_alt:
                    cell.fill = alt_fill

                # Alignment
                if isinstance(value, bool) or col_idx > 3:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # Auto-fit columns
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 45)

        ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)


def export_to_csv(rows, output_path):
    """Fallback CSV exporter."""
    if not rows:
        print("[!] No attendance records to export.")
        return
    fieldnames = list(rows[0].keys())
    with open(output_path, mode="w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main():
    parser = argparse.ArgumentParser(description="Extract PostgreSQL Detailed Attendance data to Excel (.xlsx) / CSV")
    parser.add_argument("--url", "-u", help="PostgreSQL connection URL")
    parser.add_argument("--output", "-o", default=os.getenv("OUTPUT_FILE", "detailed_attendance.xlsx"), help="Output file path")
    parser.add_argument("--format", choices=["excel", "csv"], default="excel", help="Output file format")

    args = parser.parse_args()

    # Determine URL: Argument > Environment Construct > Default Fallback
    final_url = args.url if args.url else construct_database_url()

    print("==================================================")
    print("   Detailed Attendance Exporter (By Session)      ")
    print("==================================================")

    # Hide password in logs if present
    display_url = final_url
    if "@" in display_url:
        auth_part, host_part = display_url.split("@", 1)
        scheme_part = auth_part.split("://")[0]
        display_url = f"{scheme_part}://<HIDDEN_CREDENTIALS>@{host_part}"
    print(f"[*] Database URL  : {display_url}")

    try:
        conn = psycopg2.connect(final_url, connect_timeout=4)
    except Exception as e:
        print(f"\n[!] Connection failed: {e}")
        sys.exit(1)

    sessions, attendances_by_session, all_attendance_rows = fetch_data(conn)
    conn.close()

    if not all_attendance_rows:
        print("[!] No attendance records found.")
        sys.exit(0)

    output_path = args.output

    if args.format == "csv" or not (HAS_PANDAS and HAS_OPENPYXL):
        if not output_path.endswith(".csv"):
            output_path = os.path.splitext(output_path)[0] + ".csv"
        export_to_csv(all_attendance_rows, output_path)
        print(f"[+] Exported {len(all_attendance_rows)} attendances to CSV file: {output_path}")
    else:
        if not output_path.endswith(".xlsx"):
            output_path = os.path.splitext(output_path)[0] + ".xlsx"
        export_to_excel(sessions, attendances_by_session, output_path)
        print(f"[+] Exported {len(all_attendance_rows)} attendances across {len(sessions)} sheets to Excel: {output_path}")

    print("==================================================")


if __name__ == "__main__":
    main()
